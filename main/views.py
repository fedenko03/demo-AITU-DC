import asyncio
import json
from datetime import datetime, timedelta

import qrcode
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.postgres.aggregates import ArrayAgg
from django.core.exceptions import ValidationError
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.hashers import check_password
from django.views.decorators.http import require_GET

from AITUDC import settings
from api.views import confirmed_order, canceled_order
from keytaker.consumers import WebSocketQR
from keytaker.models import Orders, History, Room, Reservation, StudyRoomSchedule, Schedule
from django.contrib import messages
from django.utils import timezone

from django.core import serializers

from keytaker.views import check_room, create_empty_cells, clear_room_schedule, fill_room_schedule
from main.models import PIN
from user.models import MainUser
import openpyxl
from openpyxl.styles import PatternFill
import pytz
from datetime import datetime
import pandas as pd
# import magic
from openpyxl import load_workbook

local_tz = pytz.timezone('Asia/Almaty')


def export_history_to_excel(request):
    # Get the queryset
    history_queryset = History.objects.all().order_by('-date')
    # Create a new workbook
    wb = openpyxl.Workbook()
    # Select the active worksheet
    ws = wb.active
    # Add headers to the worksheet
    ws['A1'] = 'Date'
    ws['B1'] = 'Fullname'
    ws['C1'] = 'Role'
    ws['D1'] = 'Room'
    ws['E1'] = 'Verified (QR)'
    ws['F1'] = 'Returned'
    # Loop over the queryset and add the data to the worksheet
    for i, entry in enumerate(history_queryset, start=2):
        ws.cell(row=i, column=1, value=entry.date.astimezone(local_tz).strftime('%d-%m-%Y %H:%M:%S'))
        ws.cell(row=i, column=2, value=entry.fullname)
        ws.cell(row=i, column=3, value=entry.role.name)
        ws.cell(row=i, column=4, value=entry.room.name)
        ws.cell(row=i, column=5, value='Yes' if entry.is_verified else 'No')
        ws.cell(row=i, column=6, value='Yes' if entry.is_return else 'No')
        # Set the background color of the row to red if is_return is True, green otherwise
        if entry.is_return:
            fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        else:
            fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        for cell in ws[i]:
            cell.fill = fill
    # Adjust the column width to fit the text
    for col in ws.columns:
        column = col[0].column_letter
        length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    # Set the filename and the content-type of the response
    now = datetime.now()
    formatted_date = now.strftime('%d.%m.%Y')
    filename = 'history ' + formatted_date + '.xlsx'
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # Create a response object
    response = HttpResponse(content_type=content_type)
    # Set the content-disposition header
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    # Write the workbook to the response
    wb.save(response)
    # Return the response
    return response


def getOrders():
    last_orders = Orders.objects.filter(
        is_available=True,
        is_confirm=False,
        orders_timestamp__gte=timezone.now() - timezone.timedelta(minutes=5)
    ).order_by('orders_timestamp')

    orders_list = []
    for order in last_orders:
        orders_list.append({
            'id': order.id,
            'room': order.room.name,
            'note': order.note,
            'user': {
                'name': order.user.full_name,
                'email': order.user.email
            },
            "orders_timestamp": order.orders_timestamp.astimezone(local_tz).strftime("%H:%M:%S"),
            "is_confirm": order.is_confirm,
            "is_available": order.is_available
        })
    return orders_list


def not_foundMain(request):
    return render(request, 'notfound404Main.html')


@login_required(login_url='loginMain')
def homeMain(request):
    orders_list = getOrders()
    link_confirm = "http://" + request.get_host() + "/"
    print(link_confirm)
    img = qrcode.make(link_confirm)
    img.save("media/mobile.png")
    return render(request, 'home-main.html', {
        'orders_list': orders_list,
        'media_url': settings.MEDIA_URL
    })


@login_required(login_url='loginMain')
def settingsMain(request):
    if request.method == "POST":
        if not request.user.is_staff:
            messages.error(request, 'Access denied')
            return redirect('loginMain')
        if 'change_password' in request.POST:
            old_password = request.POST.get('old_password')
            new_password = request.POST.get('new_password')
            checkPassword = check_password(old_password, request.user.password)
            if checkPassword:
                user_obj = User.objects.get(username=request.user)
                user_obj.set_password(new_password)
                user_obj.save()
                messages.success(request, 'Password changed successfully. Please log in again.')
                return redirect('loginMain')
            else:
                messages.error(request, 'Неверный пароль')
                return redirect('settingsMain')
        elif 'create_ec' in request.POST:
            create_empty_cells()
            messages.success(request, 'Empty cells for each room was successfully created.')
            return redirect('settingsMain')
        elif 'clear_rs' in request.POST:
            clear_room_schedule()
            messages.success(request, 'Room schedule was successfully cleared.')
            return redirect('settingsMain')
        elif 'fill_rs' in request.POST:
            fill_room_schedule()
            messages.success(request, 'Room schedule was successfully filled.')
            return redirect('settingsMain')
        elif 'upload_schedule' in request.POST:
            file = request.FILES.get('schedule_file')
            if not file:
                messages.error(request, 'Файл не найден.')
                return redirect('settingsMain')
            try:
                error = create_schedule_from_excel(file)
                if error:
                    messages.error(request, error)
                    return redirect('settingsMain')
                messages.success(request, 'Расписание успешно создано.')
                return redirect('settingsMain')
            except ValidationError as e:
                messages.error(request, 'Ошибка создания расписания.')
                return redirect('settingsMain')

    orders_list = getOrders()
    return render(request, 'settings.html', {
        'orders_list': orders_list
    })


def create_schedule_from_excel(file):
    df = pd.read_excel(file, engine='openpyxl')

    # Проверка наличия столбцов
    required_columns = ['Room', 'week_day', 'professor/email', 'start_time', 'end_time']
    missing_columns = [column for column in required_columns if column not in df.columns]

    if missing_columns:
        # Если некоторые столбцы отсутствуют, генерируем ошибку
        error_message = f"Отсутствуют необходимые столбцы: {', '.join(missing_columns)}"
        return error_message

    # Очищаем старые данные в расписании
    Schedule.objects.all().delete()

    for index, row in df.iterrows():
        room_name = row['Room']
        week_day = row['week_day']
        email = row['professor/email']
        start_time = row['start_time']
        end_time = row['end_time']

        # Проверка данных перед созданием расписания
        if not room_name or not week_day or not email or not start_time or not end_time \
                or room_name == '' or week_day == '' or email == '' or start_time == '' or end_time == '' \
                or str(room_name) == 'nan' or str(week_day) == 'nan' or str(email) == 'nan' or str(start_time) == 'nan' \
                or str(end_time) == 'nan':
            return f"Неполные данные в таблице. Проверьте все поля. ({index + 2})"

        room_obj = Room.objects.filter(name=room_name).first()
        professor_obj = MainUser.objects.filter(email=email).first()
        if not room_obj:
            return f"Проверьте все поля. Комната ' + str(room_name) + ' не найдена. ({index + 2})"
        if not professor_obj:
            return f"Проверьте все поля. Пользователь ' + str(email) + ' не найден. ({index + 2})"

        # Создание записи в расписании
        schedule = Schedule.objects.create(
            day=week_day,
            room=room_obj,
            start_time=start_time,
            end_time=end_time,
            professor=professor_obj)
        schedule.save()


def loginMain(request):
    if request.user.is_authenticated:
        return redirect('homeMain')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user_obj = User.objects.filter(username=username).first()

        if user_obj is None:
            messages.success(request, 'User not found.')
            return redirect('loginMain')
        if not user_obj.is_staff:
            messages.error(request, 'This form for admins only.')
            return redirect('loginMain')

        user = authenticate(username=username, password=password)
        if user is None:
            messages.success(request, 'Wrong password.')
            return redirect('loginMain')
        login(request, user)
        return redirect('homeMain')
    return render(request, 'login-main.html')


@login_required(login_url='loginMain')
def logoutMain(request):
    logout(request)
    return redirect('loginMain')


@login_required(login_url='loginMain')
def confirm_takeroom(request, pk):
    order_obj = Orders.objects.filter(id=pk).first()
    if not order_obj:
        messages.error(request, 'Заявка не найдена')
        return redirect('homeMain')
    error = check_room(order_obj.room.name)
    if error:
        order_obj.is_available = False
        order_obj.save()
        messages.error(request, error)
        canceled_order(order_obj.user.email, error)
        return redirect('homeMain')
    if order_obj.is_confirm or not order_obj.is_available or timezone.now() - order_obj.orders_timestamp >= timezone.timedelta(
            minutes=5):
        order_obj.is_available = False
        order_obj.save()
        canceled_order(order_obj.user.email, 'Данная заявка больше неактуальна')
        messages.error(request, 'Данная заявка больше неактуальна')
        return redirect('homeMain')

    if request.method == 'POST':
        code = request.POST.get('code')
        order_obj = Orders.objects.filter(confirmation_code=code).first()

        if not order_obj:
            messages.error(request, 'Неверный код или заявка')
            return redirect('confirm-takeroom', pk)

        order_obj.is_available = False
        order_obj.is_confirm = True
        history = History.objects.create(
            room=order_obj.room,
            fullname=order_obj.user.full_name,
            is_verified=True,
            role=order_obj.user.role,
            user=order_obj.user,
            date=timezone.now()
        )
        room_obj = Room.objects.filter(name=order_obj.room.name).first()
        room_obj.is_occupied = True
        room_obj.save()
        history.save()
        order_obj.save()
        confirmed_order(order_obj.user.email, 'Заявка подтверждена успешно')
        messages.success(request, 'Заявка подтверждена успешно')
        return redirect('homeMain')
    else:
        orders_list = {'id': pk,
                       'room': order_obj.room.name,
                       'note': order_obj.note,
                       'user': {
                           'name': order_obj.user.full_name,
                           'email': order_obj.user.email
                       },
                       "orders_timestamp": order_obj.orders_timestamp,
                       "is_confirm": order_obj.is_confirm
                       }
        print(orders_list)
        orders1_list = getOrders()
        return render(request, 'confirm-takeroom-main.html', {
            'orders_list': orders1_list,
            'order': orders_list
        })


def cancel_takeroomMain(request, pk):
    order_obj = Orders.objects.filter(id=pk).first()
    if order_obj:
        order_obj.is_available = False
        order_obj.save()
        msg = 'Заявка успешно отклонена.'
        canceled_order(order_obj.user.email, 'Заявка была отклонена. Попробуйте снова')
        messages.error(request, msg)
        return redirect('homeMain')
    return redirect('confirm-takeroom', pk)


@login_required(login_url='loginMain')
def historyMain(request):
    history_entries = History.objects.all().order_by('-date')
    orders_list = getOrders()
    return render(request, 'history-main.html', {
        'orders_list': orders_list,
        'history_entries': history_entries,
        'history_length': len(history_entries)
    })


def history_ajax(request):
    start = int(request.GET.get('start', 0))
    end = int(request.GET.get('end', 10))
    search_query = request.GET.get('search_query')
    history_length = 0
    if search_query:
        # print(1)
        room_obj = Room.objects.filter(name=search_query).first()
        if room_obj:
            history = History.objects.filter(room=room_obj).order_by('-date')
            history_length = len(history)
            history = history[start:end]
        else:
            start = 0
            end = 0
            history = []
    else:
        # print(2)
        history = History.objects.all().order_by('-date')
        history_length = len(history)
        history = history[start:end]
    history_data = []
    for entry in history:
        history_data.append({
            'fullname': entry.fullname,
            'role': entry.role.name,
            'room': entry.room.name,
            'date': entry.date.astimezone(local_tz).strftime('%d-%m-%Y %H:%M:%S'),
            'is_verified': entry.is_verified,
            'is_return': entry.is_return,
        })
    # print(history_data)
    # print(history_length)
    data = {
        'history_json': history_data,
        'history_length': history_length,
        'start': start,
        'end': end
    }
    return JsonResponse(data)


@login_required(login_url='loginMain')
def usersMain(request):
    query = request.GET.get('q')
    if query:
        users_obj = MainUser.objects.filter(full_name__icontains=query)
    else:
        users_obj = MainUser.objects.all().order_by('full_name')
    orders_list = getOrders()
    context = {
        'orders_list': orders_list,
        'users_obj': users_obj
    }
    return render(request, 'users-main.html', context)


@login_required(login_url='loginMain')
def roomsMain(request):
    query = request.GET.get('q')
    orders_list = getOrders()
    if query:
        rooms_list = Room.objects.filter(name=query)
    else:
        rooms_list = Room.objects.filter(floor='1').all().order_by('name')
    return render(request, 'rooms-main.html', {
        'orders_list': orders_list,
        'rooms_obj': rooms_list
    })


@login_required(login_url='loginMain')
def confirmBooking(request, key):
    try:
        reservation_obj = Reservation.objects.get(key=key)
    except Reservation.DoesNotExist:
        messages.error(request, "Reservation not found")
        return redirect('homeMain')

    studyroom_schedule = StudyRoomSchedule.objects.filter(
        room=reservation_obj.room,
        start_time=reservation_obj.start_time
    ).first()

    if not studyroom_schedule:
        messages.error(request, 'The room ' + reservation_obj.room.name + ' does not have a schedule or this time slot.')
        return redirect('homeMain')
    if studyroom_schedule.status != 'free':
        messages.error(request, "The selected time slot is not available.")
        return redirect('homeMain')

    # Check if the reservation is still valid (created less than 5 minutes ago)
    time_diff = timezone.now() - reservation_obj.created_at
    if time_diff > timezone.timedelta(minutes=5):
        messages.error(request, "The reservation has expired.")
        return redirect('homeMain')

    # Check if the reservation is not yet activated
    if reservation_obj.is_active:
        messages.error(request, "The reservation has already been confirmed.")
        return redirect('homeMain')

    link_confirm = "http://" + request.get_host() + "/reserve-studyroom/" + reservation_obj.key + "/"
    print(link_confirm)
    img = qrcode.make(link_confirm)
    img.save("media/bookingQR.png")

    status_list = []
    status_list.append({'notification_type': 'key_booking',
                        'data': {
                            'link_confirm': link_confirm,
                            'qr_url': settings.MEDIA_URL + 'bookingQR.png',
                            'timestamp': (reservation_obj.created_at + timedelta(minutes=5)).astimezone(local_tz).strftime("%H:%M:%S %d.%m.%Y"),
                            'room': reservation_obj.room.name,
                            'time': reservation_obj.start_time.strftime("%H"),
                            'is_take': reservation_obj.is_take
                        }})
    for consumer in WebSocketQR.consumers:
        asyncio.run(consumer.send(text_data=json.dumps(status_list)))

    return render(request, 'confirm-booking-qr.html', {
        'media_url': settings.MEDIA_URL,
        'room': reservation_obj.room.name,
        'start_time': reservation_obj.start_time.strftime('%H:%M'),
        'key': reservation_obj.key,
        'is_take': reservation_obj.is_take,
        'is_active': reservation_obj.is_active
    })


def get_rooms_floor(request):
    floor = request.GET.get('floor')
    rooms_list = Room.objects.filter(floor=floor).annotate(role_name_list=ArrayAgg('role__name')).values('name',
                                                                                                         'description',
                                                                                                         'is_occupied',
                                                                                                         'is_visible',
                                                                                                         'map_id',
                                                                                                         'role_name_list').order_by(
        'name')
    if not rooms_list:
        return JsonResponse({
            'rooms_list': None
        })
    return JsonResponse({
        'rooms_list': list(rooms_list)
    })


def get_room_map(request):
    room_map_id = request.GET.get('room_map_id')
    room_obj = Room.objects.filter(map_id=room_map_id).first()

    if not room_obj:
        return JsonResponse({
            'room_map_info': None
        })

    user_fullname = ''
    if room_obj.is_occupied:
        history_obj = History.objects.filter(
            room=room_obj,
            is_return=False
        ).first()
        if not history_obj:
            return JsonResponse({
                'room_map_info': None
            })
        print(history_obj)
        user_fullname = history_obj.fullname

    roles = [role.name for role in room_obj.role.all()]
    room_map_info = {
        'map_id': room_obj.map_id,
        'name': room_obj.name,
        'role': roles,
        'description': room_obj.description,
        'is_occupied': room_obj.is_occupied,
        'is_visible': room_obj.is_visible,
        'user_fullname': user_fullname
    }
    # print(room_map_info)
    return JsonResponse({'room_map_info': room_map_info})


@login_required(login_url='loginMain')
def pinLocked(request):
    if request.method == "POST":
        code = request.POST.get("pincode")
        code_obj = PIN.objects.filter(code=code).first()
        if not code_obj:
            messages.error(request, 'Неверный PIN-код')
            return redirect('pinLocked')
        code_obj.is_locked = False
        code_obj.save()
        return redirect('homeMain')
    return render(request, 'pin.html')


@login_required(login_url='loginMain')
def websocketQR(request):
    return render(request, 'websocket-qr.html', {
        'media_url': settings.MEDIA_URL,
    })


@login_required(login_url='loginMain')
def PinLock(request, code):
    pinLock_obj = PIN.objects.first()
    if pinLock_obj.is_locked:
        messages.error(request, 'Доступ был заблокирован ранее')
        return redirect('homeMain')

    if code.isdigit() and 4 <= len(code) <= 10:
        pinLock_obj.is_locked = True
        pinLock_obj.code = code
        pinLock_obj.save()
        return redirect('pinLocked')
    else:
        messages.error(request, 'PIN-код должен содержать от 4 до 10 цифр.')
        return redirect('homeMain')
